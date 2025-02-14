# include libraries :
import openpyxl as excel
from openpyxl.chart import LineChart, Reference
from customtkinter import filedialog as fd
from datetime import datetime 
import time
import matplotlib.pyplot as plt
import os

def getDictionary(records : list, values : list, time : list, sub_time : list, date : list) -> dict:
    """
        this function return the dictionary as (time seconds : value in this second)
    """
    # convert the lists to dictionary and return it
    return {
    str(key): {
        "value": value,
        "time": t,
        "sub_time": st,
        "date": d
    } for key, value, t, st, d in zip(records, values, time, sub_time, date)}
def saveGraphHighResolution(records, values, file_path):
    # Create a figure and axis
    fig, ax = plt.subplots()

    # Plot the data
    ax.plot(records, values, linestyle='-')

    # Add labels and title
    ax.set_ylabel('Values')
    ax.set_xlabel('Records')
    ax.set_title('Sensor data')

    # Save the figure in SVG format for high-quality scaling
    plt.savefig(f'{file_path} - {datetime.now().date().strftime("%d-%m-%Y")}.svg', format='svg')

    # Alternatively, you can save as PDF
    plt.savefig(f'{file_path} - {datetime.now().date().strftime("%d-%m-%Y")}.pdf', format='pdf')
def showGraph(records, values):
    # Create a figure and axis
    fig, ax = plt.subplots()

    # Plot the data
    ax.plot(records, values, linestyle='-')

    # Add labels and title
    ax.set_ylabel('Values')
    ax.set_xlabel('Records')
    ax.set_title('Sensor Value')

    # Show the plot
    plt.show()

def exportExcelSheet(record_name : str, dictionary: dict, records : list, values : list, port_name : str, baud_rate : str, start_time_string : str, len_records : int, date : datetime, end_time_string : str, duration : time):
    """
        this function has created to inset data into excel sheet and save it
    """

    # create object to open the exel sheet :
    paper = excel.Workbook()

    # open the exel sheet :
    sheet = paper.active

    # title of exel sheet :
    sheet.title = "Data Value"

    # # details :

    # title name of the sheet :
    title_cell = sheet.cell(row=1, column=3)
    # value of the title sheet:
    title_cell.value = f"{record_name}"

    # start time of recording :
    start_time_cell = sheet.cell(row=2, column=1)
    # value of the start value:
    start_time_value_cell = sheet.cell(row=2, column=2)

    # cells of them :
    start_time_cell.value = "Start Time"
    start_time_value_cell.value = f"{start_time_string}"

    # end time of recording :
    end_time_cell = sheet.cell(row=2, column=4)
    # value of the end value:
    end_time_value_cell = sheet.cell(row=2, column=5)

    # cells of them:
    end_time_cell.value = "End Time"
    end_time_value_cell.value = f"{end_time_string}"

    # data of recording : 
    date_cell = sheet.cell(row=3, column=1)
    # data value of recording :
    date_value_cell = sheet.cell(row=3, column=2)

    # cells :
    date_cell.value = "Date"
    date_value_cell.value = f"{date}"
    
    # duration of recording :
    duration_cell = sheet.cell(row=4, column=1)
    # duration value of the cell:
    duration_value_cell = sheet.cell(row=4, column=2)

    # cells :
    duration_cell.value = "Duration"    
    duration_value_cell.value = f"{duration}"

    # records by time :
    records_cell = sheet.cell(row=5, column=1)
    # records value by time :
    records_value_cell = sheet.cell(row=5, column=2)

    # cells:
    records_cell.value = "Records"
    records_value_cell.value = f"{len_records}"

    # records by time :
    port_cell = sheet.cell(row=6, column=1)
    # records value by time :
    port_value_cell = sheet.cell(row=6, column=2)

    # cells:
    port_cell.value = "Port name"
    port_value_cell.value = f"{port_name}"


    # baud rate :
    baud_rate_cell = sheet.cell(row=7, column=1)
    # baud rate value :
    baud_rate_value_cell = sheet.cell(row=7, column=2)

    # cells:
    baud_rate_cell.value = "Baud rate"
    baud_rate_value_cell.value = f"{baud_rate}"


    #  select the time, value, tim every record, sub time, date (labels) cell:
    time_cell = sheet.cell(row=9, column=1)
    value_cell = sheet.cell(row=9, column=2)
    time_str_cell = sheet.cell(row=9, column=3)
    sub_time_cell = sheet.cell(row=9, column=4)
    date_str_cell = sheet.cell(row=9, column=5)

    # insert values into cells
    time_cell.value = "Records"
    value_cell.value = "Values"
    time_str_cell.value = "Time"
    sub_time_cell.value = "Sub time"
    date_str_cell.value = "Date"

    # init the row number to start from second row
    row_number = 10

    for key, value in dictionary.items():
        #  select the time, value cell:
        time_cell = sheet.cell(row=row_number, column=1)
        value_cell = sheet.cell(row=row_number, column=2)
        time_str_cell = sheet.cell(row=row_number, column=3)
        sub_time_cell = sheet.cell(row=row_number, column=4)
        date_str_cell = sheet.cell(row=row_number, column=5)

        # insert values into cells
        time_cell.value = key
        value_cell.value = value["value"]
        time_str_cell.value = value["time"]
        sub_time_cell.value = value["sub_time"]
        date_str_cell.value = value["date"]

        #  increase the row number 
        row_number += 1

    # Create a LineChart object
    chart = LineChart()
    chart.title = "Sensor Value Reading"
    chart.style = 13
    chart.x_axis.title = 'Seconds'
    chart.y_axis.title = 'Values'

    # Define the data for the chart
    data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=len(dictionary) + 9)
    chart.add_data(data, titles_from_data=True)

    # Define the categories (x-axis labels)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=len(dictionary) + 9)
    chart.set_categories(cats)

    # Add the chart to the worksheet
    sheet.add_chart(chart, "E5")

    # aks where do you like to save the exel file
    file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Exel File", ".xlsx")])

    # get only path t make dir .
    path = file_path[:len(file_path)-5]

    # get only name the user entered it.
    index = file_path.rfind('/')
    file_name = file_path[index +1 :len(file_path)-5]

    # make file .
    os.makedirs(f"{path}", exist_ok=True)
    
    # export high resolution graph :
    saveGraphHighResolution(records, values, f"{path}\{file_name}")

    # save the file
    paper.save(f"{path}\{file_name}.xlsx")


